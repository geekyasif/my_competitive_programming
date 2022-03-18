# # importing openpyxl module
# import openpyxl
# import pandas as pd
# import xlrd
# from sqlalchemy import create_engine
# import mysql.connector
# import os
# conn = mysql.connector.connect(host="localhost", user="root", passwd="", database="excel")
# cur = conn.cursor()
#
# # Give the location of the file
# path = "C:\\Users\\Mohammad Asif\\Downloads\\FINAL450.xlsx"
# import openpyxl, pprint, json
# print('Opening workbook...')

# wb = xlrd.open_workbook(path)
# sheet = wb.sheet_by_index(0)
# sheet.cell_value(0, 0)
#
# for i in range(sheet.nrows):
#     print(sheet.cell_value(i, 0))

# # workbook object is created
# wb_obj = openpyxl.load_workbook(path)
#
# sheet_obj = wb_obj.active
# m_row = sheet_obj.max_column
# print(m_row)
# # print(sheet_obj.values)
#
# # Loop will print all values
# # of first column
# for i in range(1, m_row + 1):
#     cell_obj = sheet_obj.cell(row=i, column=i)
#     print(cell_obj.value, end=",")
# #
# json = excel2json.convert_from_file(path)
# print(json)
# xls = pd.ExcelFile(path)
# df1 = pd.read_excel(path)
# json = df1.to_json()
# print(json)
# print(df1)


# conn.close()
#
# a = xlrd.open_workbook(path)
# sheet = a.sheet_by_index(0)
# sheet.cell_value(0,0)
# for i in range(1,20):
#     print(sheet.row_values(i))


# import excel2json
#
# json = excel2json.convert_from_file(path)
#
# print(json)

# book = xlrd.open_workbook(path)
# sheet = book.sheet_by_name("Sheet1")
# print(sheet)

# workbook object is created
# wb_obj = openpyxl.load_workbook(path)
#
# sheet_obj = wb_obj.active
# print(sheet_obj['A1'])
#
# max_col = sheet_obj.max_column
#
# # Will print a particular row value
# for i in range(1, max_col + 1):
#     cell_obj = sheet_obj.cell(row=i, column=i)
#     print(cell_obj.value, end=" ")

import openpyxl, pprint, json
print('Opening workbook...')
wb = openpyxl.load_workbook('excel_form.xlsx')
sheet = wb.get_sheet_by_name('Sheet')


excel_data = {}
print('Reading rows...')
for row in range(2, sheet.max_row + 1):
    Col F  = sheet['F' + str(row)].value
    Col C = sheet['C' + str(row)].value
    Col A = sheet['A' + str(row)].value

    excel_data.setdefault(Col F, {})
    excel_data[Col F].setdefault(Col C, {'Col A': Col A})


# Open a new text file and write the contents of excel_data to it.
print('Writing results...')
with open('DATA.json', 'w') as resultFile:
    json.dump(Matrix, resultFile)
print('Done.')